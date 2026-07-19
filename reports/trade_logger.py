"""Exports the trade log to CSV and a formatted Excel workbook."""
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import config

TRADE_FIELDS = [
    "date", "entry_time", "exit_time", "trade_type", "strike", "option_symbol",
    "entry_premium", "exit_premium", "stop_loss", "target",
    "pnl_points", "pnl_rupees", "exit_reason", "duration"
]


def trade_to_row(t):
    return {
        "date": t.date,
        "entry_time": t.entry_time.strftime("%H:%M:%S") if hasattr(t.entry_time, "strftime") else str(t.entry_time),
        "exit_time": t.exit_time.strftime("%H:%M:%S") if (t.exit_time and hasattr(t.exit_time, "strftime")) else (str(t.exit_time) if t.exit_time else ""),
        "trade_type": t.trade_type,
        "strike": t.strike,
        "option_symbol": t.option_symbol,
        "entry_premium": t.entry_premium,
        "exit_premium": t.exit_premium,
        "stop_loss": t.stop_loss,
        "target": t.target,
        "pnl_points": t.pnl_points(),
        "pnl_rupees": t.pnl_rupees(),
        "exit_reason": t.exit_reason,
        "duration": t.duration(),
    }


def write_csv(trades, path=None):
    path = path or os.path.join(config.EXPORTS_DIR, "trade_log.csv")
    rows = [trade_to_row(t) for t in trades]
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TRADE_FIELDS)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    return rows, path


def write_xlsx(rows, eod_stats, path=None):
    path = path or os.path.join(config.EXPORTS_DIR, "trade_log.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trade Log"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    normal_font = Font(name="Arial")

    headers = ["Date", "Entry Time", "Exit Time", "Type", "Strike", "Symbol",
               "Entry Premium", "Exit Premium", "Stop Loss", "Target",
               "P/L (Points)", "P/L (Rs)", "Exit Reason", "Duration"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([r["date"], r["entry_time"], r["exit_time"], r["trade_type"], r["strike"],
                   r["option_symbol"], r["entry_premium"], r["exit_premium"], r["stop_loss"],
                   r["target"], r["pnl_points"], r["pnl_rupees"], r["exit_reason"], r["duration"]])

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(length + 2, 10)

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = normal_font

    green = Font(name="Arial", color="006100")
    red = Font(name="Arial", color="9C0006")
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.font = green if c.value >= 0 else red

    ws2 = wb.create_sheet("EOD Report")
    ws2.append(["Metric", "Value"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
    for k, v in eod_stats.items():
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20
    for row in ws2.iter_rows(min_row=2):
        row[0].font = normal_font
        row[1].font = normal_font

    wb.save(path)
    return path
